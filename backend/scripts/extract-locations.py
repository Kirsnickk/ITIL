#!/usr/bin/env python3
"""
Extract real location data from Excel with proper schema fields.
Generates locations.json with: code, name, type, address, city, country, status
"""

import openpyxl
import json
from pathlib import Path

def extract_locations():
    """Extract locations from Store_Info_VN sheet"""
    
    excel_path = Path(__file__).parent.parent.parent / "Data" / "Infra Management.xlsx"
    output_path = Path(__file__).parent / "extracted-data" / "locations.json"
    
    print(f"📂 Reading Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    locations = []
    
    # Extract STORES from Store_Info_VN sheet
    print("\n🏪 Extracting stores from Store_Info_VN...")
    if 'Store_Info_VN' in wb.sheetnames:
        sheet = wb['Store_Info_VN']
        headers = [str(cell.value) for cell in sheet[1] if cell.value]
        
        store_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            store_data = {headers[i]: val for i, val in enumerate(row) if i < len(headers)}
            
            # Skip if no store code
            if not store_data.get('Store code'):
                continue
            
            # Map status
            status = 'ACTIVE'
            if store_data.get('Status'):
                status_val = str(store_data['Status']).strip().lower()
                if status_val in ['close', 'closed', 'inactive']:
                    status = 'INACTIVE'
            
            location = {
                'code': str(store_data['Store code']).strip(),
                'name': str(store_data.get('Store Name', store_data['Store code'])).strip(),
                'type': 'STORE',
                'address': str(store_data.get('Address', '')).strip() or f"Store {store_data['Store code']}",
                'city': str(store_data.get('City', 'Ho Chi Minh City')).strip(),
                'country': 'Vietnam',
                'status': status,
            }
            
            # Optional fields
            if store_data.get('Email'):
                location['email'] = str(store_data['Email']).strip()
            
            locations.append(location)
            store_count += 1
        
        print(f"   ✓ Extracted {store_count} stores")
    
    # Add HEAD OFFICE location (always needed)
    head_office = {
        'code': 'HO-VN',
        'name': 'Head Office - Ho Chi Minh',
        'type': 'OFFICE',
        'address': 'Ho Chi Minh City, Vietnam',
        'city': 'Ho Chi Minh City',
        'country': 'Vietnam',
        'status': 'ACTIVE',
    }
    locations.append(head_office)
    print(f"   ✓ Added Head Office")
    
    # Save to JSON
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(locations, f, indent=2, ensure_ascii=False)
    
    print(f"\n✅ Saved {len(locations)} locations to: {output_path}")
    print(f"\n📊 Summary:")
    print(f"   - Stores: {sum(1 for loc in locations if loc['type'] == 'STORE')}")
    print(f"   - Offices: {sum(1 for loc in locations if loc['type'] == 'OFFICE')}")
    print(f"   - Active: {sum(1 for loc in locations if loc['status'] == 'ACTIVE')}")
    print(f"   - Inactive: {sum(1 for loc in locations if loc['status'] == 'INACTIVE')}")
    
    return locations

if __name__ == '__main__':
    locations = extract_locations()
    print("\n🎉 Location extraction complete!")
